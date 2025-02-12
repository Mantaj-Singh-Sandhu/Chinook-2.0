import os
import pandas as pd
import re
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import sys

# Ensure required packages are installed
try:
    import pandas as pd
    from sqlalchemy import create_engine
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Required package not found: {e}")
    sys.exit(1)

# Load input file name from a text file
input_folder = 'parquet'
input_filename_file = 'input_file.txt'

try:
    with open(input_filename_file, 'r') as f:
        input_file = f.read().strip()
    if not input_file:
        raise ValueError("Input filename is empty.")
except Exception as e:
    print(f"Error reading the input filename: {e}")
    sys.exit(1)

input_path = os.path.join(input_folder, input_file)

# Load input data from Parquet file
try:
    df = pd.read_parquet(input_path)
    if df.empty:
        print("Input file is empty, proceeding to create files anyway.")
except Exception as e:
    print(f"Error loading the input file: {e}")
    sys.exit(1)

# Ensure 'name' column exists
if 'name' not in df.columns:
    print("Error: 'name' column is missing from the DataFrame!")
    sys.exit(1)

# Function to remove illegal characters
def remove_illegal_characters(value):
    if isinstance(value, str):
        return ''.join(c for c in value if c.isprintable())
    return value

# Apply cleaning only to string columns
for col in df.select_dtypes(include=['object']).columns:
    df[col] = df[col].apply(remove_illegal_characters)

# Create output folder if it doesn't exist
output_folder = 'excel_outputs'
os.makedirs(output_folder, exist_ok=True)

# Function to apply colors, styles, and column widths to the Excel file
def apply_excel_formatting(excel_file_path):
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Define the header color (light blue) and black font
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=12)

    # Apply header color and bold, black font with increased font size
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Dynamically adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(50, max(max_length + 2, 10))  # Limit max width to 50

    # Save the formatted workbook
    wb.save(excel_file_path)

# Function for filtering and saving with applied formatting
def filter_and_save(df, filter_condition, output_filename, exclude_columns=None):
    print(f"Processing filter for {output_filename}...")
    try:
        filtered_df = df[filter_condition].copy()
        print(f"Rows matching filter for {output_filename}: {len(filtered_df)}")

        # If no data matches, create a file with just headers
        if filtered_df.empty:
            print(f"No data to save for {output_filename}. Creating an empty file with headers.")
            filtered_df = pd.DataFrame(columns=['name', 'value', 'duplicate_count'])
            filtered_df.to_excel(os.path.join(output_folder, output_filename), index=False)
            apply_excel_formatting(os.path.join(output_folder, output_filename))
            print(f"✅ Successfully saved empty {output_filename}")
            return

        if exclude_columns:
            filtered_df = filtered_df[~filtered_df['name'].str.contains('|'.join(exclude_columns), case=False, na=False)]

        filtered_df['duplicate_count'] = filtered_df.groupby(['name', 'value'])['name'].transform('count')
        filtered_unique = filtered_df.drop_duplicates(subset=['name', 'value'])
        filtered_unique = filtered_unique[['name', 'value', 'duplicate_count']]
        filtered_unique = filtered_unique.sort_values(by='name', ascending=True)

        for col in filtered_unique.select_dtypes(include=['object']).columns:
            filtered_unique[col] = filtered_unique[col].apply(remove_illegal_characters)

        output_file_path = os.path.join(output_folder, output_filename)
        filtered_unique.to_excel(output_file_path, index=False)
        apply_excel_formatting(output_file_path)
        print(f"✅ Successfully saved {output_filename}")
    except Exception as e:
        print(f"❌ Error processing {output_filename}: {e}")

# Define filters for processing
filters = {
    'athena_query_results_dtc_CDL_with_count.xlsx': {'include': ['DTC'], 'exclude': None},  # Special case handled separately
    'athena_query_results_dtc_J1939_with_count.xlsx': {'include': ['DTC'], 'exclude': None},  # Special case handled separately
    'athena_query_results_OoR_CDL_with_count.xlsx': {'include': ['OoR'], 'exclude': None},  # Special case handled separately
    'athena_query_results_OoR_J1939_with_count.xlsx': {'include': ['OoR'], 'exclude': None},  # Special case handled separately
    'athena_query_results_fmi.xlsx': {'include': ['^CDLECM'], 'exclude': None},
    'athena_query_results_LAMP.xlsx': {'include': ['LAMP'], 'exclude': None},
    'athena_query_results_CDLWarning.xlsx': {'include': ['^CDLWarning'], 'exclude': None},
    'athena_query_results_DM1_DM2_no_duplicates.xlsx': {'include': ['DM1','DM2'], 'exclude': None},
    'athena_query_results_error_no_duplicates.xlsx': {'include': ['error'], 'exclude': None},
    'athena_query_results_j1939_no_error_dtc_rpm_with_count.xlsx': {'include': ['J1939'], 'exclude': ['error', 'DTC', 'DM1', 'DM2', 'RPM']},
    'athena_query_results_rpm_with_count.xlsx': {'include': ['RPM'], 'exclude': None},
    'athena_query_results_cdl_no_dtc_error_rpm_cdlecm_with_count.xlsx': {'include': ['CDL'], 'exclude': ['DTC', 'error', 'RPM', 'CDLECM*']},
}

# Apply filtering and saving for each filter
for filename, conditions in filters.items():
    include = conditions['include']
    exclude = conditions['exclude']

    if filename == "athena_query_results_dtc_CDL_with_count.xlsx":
        filter_condition = df['name'].str.startswith('CDL', na=False) & df['name'].str.endswith('DTC', na=False)

    elif filename == "athena_query_results_fmi.xlsx":
        # Ensure 'name' starts with "CDLECM"
        filter_condition = df['name'].str.startswith("CDLECM", na=False)

    elif filename == "athena_query_results_CDLWarning.xlsx":
        # Ensure 'name' starts with "CDLECM"
        filter_condition = df['name'].str.startswith("CDLWarning", na=False)

    elif filename == "athena_query_results_dtc_J1939_with_count.xlsx":
        filter_condition = df['name'].str.startswith('J1939', na=False) & df['name'].str.endswith('DTC', na=False)

    elif filename == "athena_query_results_OoR_CDL_with_count.xlsx":
        filter_condition = df['name'].str.startswith('CDL', na=False) & df['name'].str.endswith('OoR', na=False)

    elif filename == "athena_query_results_OoR_J1939_with_count.xlsx":
        filter_condition = df['name'].str.startswith('J1939', na=False) & df['name'].str.endswith('OoR', na=False)       

    else:
        filter_condition = df['name'].str.contains('|'.join(map(re.escape, include)), case=False, na=False)

    filter_and_save(df, filter_condition, filename, exclude_columns=exclude)
