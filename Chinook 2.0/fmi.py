import pandas as pd
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Paths to input and output files
input_file = "excel_outputs/athena_query_results_fmi.xlsx"  # Input file path
fmi_source_file = "FMISource.xlsx"  # FMI source file in the same folder as the script
output_file = "excel_outputs/FMI-CID.xlsx"  # Output file path

# Ensure the output file is created (even if no data to write)
empty_df = pd.DataFrame(columns=["CID Description", "FMI Description", "count", "fmi", "cid", "active"])
empty_df.to_excel(output_file, index=False)

try:
    # Step 1: Read the input Excel file
    print("Reading the Excel file...")
    input_df = pd.read_excel(input_file, sheet_name=None)  # Read all sheets to dynamically handle the sheet
    sheet_name = list(input_df.keys())[0]  # Automatically get the first sheet name
    input_df = input_df[sheet_name]

    # Step 2: Parse JSON data in the 'value' column
    print("Parsing JSON data...")
    parsed_data = []
    for _, row in input_df.iterrows():
        try:
            json_data = json.loads(row['value'])  # Assuming the JSON is stored in the 'value' column
            for entry in json_data:
                parsed_data.append({
                    "fmi": entry.get("fmi", None),  # Default to None if key is missing
                    "cid": entry.get("cid", None),
                    "active": entry.get("active", None)
                })
        except (json.JSONDecodeError, KeyError, TypeError):
            print(f"Skipping row due to error: {row.get('value', None)}")

    # Convert parsed data to a DataFrame
    parsed_df = pd.DataFrame(parsed_data)
    if parsed_df.empty:
        print("No valid data found in the JSON parsing step.")

    # Step 3: Remove duplicates and count them
    print("Counting duplicates...")
    if 'fmi' in parsed_df.columns and 'cid' in parsed_df.columns:
        # Add a 'count' column with default value of 1
        parsed_df['count'] = 1
        if not parsed_df.empty:
            # Update 'count' with actual duplicate counts
            parsed_df['count'] = parsed_df.groupby(["fmi", "cid", "active"])["fmi"].transform("count")
        # Ensure duplicate removal but keep the 'count' column
        parsed_df = parsed_df.drop_duplicates(["fmi", "cid", "active"]).reset_index(drop=True)
    else:
        print("Required columns 'fmi' or 'cid' are missing. Exiting.")

    # Step 4: Read the FMI Source file
    print("Reading the FMI Source data...")
    fmi_df = pd.read_excel(fmi_source_file, sheet_name="FMI")
    fmi_df.columns = fmi_df.columns.str.strip().str.lower()  # Clean column names
    fmi_df['fmi'] = fmi_df['fmi'].apply(lambda x: str(x).replace('\xa0', ' ').strip())  # Clean FMI column

    # Step 5: Add description based on FMI
    print("Adding descriptions based on fmi...")
    parsed_df["FMI Description"] = [
        fmi_df.loc[fmi_df["fmi"] == str(fmi).strip(), "description"].values[0].strip()
        if len(fmi_df.loc[fmi_df["fmi"] == str(fmi).strip(), "description"].values) > 0 else "No Description"
        for fmi in parsed_df["fmi"]
    ]

    # Step 6: Read the CID descriptions from the "CID" sheet in FMISource
    print("Reading the CID descriptions...")
    cid_df = pd.read_excel(fmi_source_file, sheet_name="CID")
    cid_df.columns = cid_df.columns.str.strip().str.lower()  # Clean column names
    cid_df['cid'] = cid_df['cid'].apply(lambda x: str(x).replace('\xa0', ' ').strip())  # Clean CID column

    # Step 7: Add CID description based on CID number
    print("Adding CID descriptions based on cid...")
    parsed_df["CID Description"] = [
        cid_df.loc[cid_df["cid"] == str(cid).strip(), "description"].values[0].strip()
        if len(cid_df.loc[cid_df["cid"] == str(cid).strip(), "description"].values) > 0 else "No Description"
        for cid in parsed_df["cid"]
    ]

    # Step 8: Reorder columns: CID Description first, then FMI Description
    parsed_df = parsed_df[["CID Description", "FMI Description", "count", "fmi", "cid", "active"]]

    # Step 9: Save the parsed DataFrame to a new Excel file
    print("Writing output to Excel...")
    parsed_df.to_excel(output_file, index=False)

    # Step 10: Adjust column widths and apply color formatting using openpyxl
    print("Adjusting column widths and applying color formatting...")
    wb = load_workbook(output_file)
    ws = wb.active

    # Define the desired column widths
    column_widths = {
        'A': 40,  # CID Description column
        'B': 40,  # FMI Description column
        'C': 10,  # count column
        'D': 10,  # fmi column
        'E': 10,  # cid column
        'F': 10   # active column
    }

    # Adjust each column's width
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Apply a light blue color to the first row (header)
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    for cell in ws[1]:
        cell.fill = light_blue_fill

    # Save the adjusted file
    wb.save(output_file)
    print(f"Conversion complete! Data saved to {output_file}")

    # Step 11: Remove the input file
    os.remove(input_file)
    print(f"Input file '{input_file}' has been removed.")

except Exception as e:
    print(f"An error occurred: {e}")
