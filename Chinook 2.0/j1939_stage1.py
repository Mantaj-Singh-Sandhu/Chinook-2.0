import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os  # Ensure this is imported

# File paths
input_file = "excel_outputs/athena_query_results_j1939_no_error_dtc_rpm_with_count.xlsx"
athena_file = "excel_outputs/Format_temp.xlsx"
j1939_limits_file = "j1939_limit.xlsx"

# Step 1: Read the input and reference Excel files
aws_df = pd.read_excel(input_file)
limits_df = pd.read_excel(j1939_limits_file)

# Step 2: Standardize column names
aws_df.columns = aws_df.columns.str.strip()
limits_df.columns = limits_df.columns.str.strip()

# Step 3: Round the 'value' column to remove floating-point discrepancies and convert to integer
aws_df['value'] = pd.to_numeric(aws_df['value'], errors='coerce')
aws_df['value'] = aws_df['value'].fillna(0).astype(int)
aws_df['value'] = aws_df['value'].round().astype(int)
aws_df['value'] = aws_df['value'].round().astype(int)

# Step 4: Group by 'name' and 'value', summing 'duplicate_count' and removing duplicates
# Before grouping, let's print the duplicates
duplicates = aws_df[aws_df.duplicated(subset=['name', 'value'], keep=False)]

if not duplicates.empty:
    print("Duplicates before combining:")
    print(duplicates)

# Grouping and summing duplicate_count
aws_df = aws_df.groupby(['name', 'value'], as_index=False)['duplicate_count'].sum()

# Step 5: Reorder columns for output
aws_df = aws_df[['name', 'value', 'duplicate_count']]

# Step 6: Sort based on the sequence in `j1939_limits.xlsx` and move unmatched tags to the end
sorted_names = limits_df['name'].dropna().tolist()

# Create a function to return a sort order, putting unmatched tags at the end
def get_sort_order(tag_name):
    if tag_name in sorted_names:
        return sorted_names.index(tag_name)
    else:
        return len(sorted_names)  # Put unmatched tags at the end

# Apply the sorting logic and reorder
aws_df['name_order'] = aws_df['name'].apply(get_sort_order)
aws_df = aws_df.sort_values(by='name_order').drop(columns=['name_order'])

# Step 7: Load the input Excel file using openpyxl
workbook = load_workbook(input_file)
sheet = workbook.active

# Define fills for formatting
blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# Step 8: Preserve existing red marks
existing_fills = {}
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):  # Only preserve 'name' and 'value' fills
    for cell in row:
        if cell.fill.start_color.index != "00000000":  # Check if cell already has a fill
            existing_fills[(cell.row, cell.column)] = (cell.fill.start_color.index, cell.fill.end_color.index)

# Step 9: Clear and rewrite the sheet with formatted data
sheet.delete_rows(2, sheet.max_row)  # Remove all rows except the header

# Write the header
header = ['name', 'value', 'duplicate_count']  # Reordered columns
for col_idx, header_name in enumerate(header, start=1):
    header_cell = sheet.cell(row=1, column=col_idx, value=header_name)
    header_cell.fill = blue_fill  # Apply light blue fill to the header

# Write the new data
for row_idx, row in enumerate(aws_df.itertuples(index=False), start=2):
    sheet.cell(row=row_idx, column=1, value=row.name)  # Name column
    sheet.cell(row=row_idx, column=2, value=row.value)  # Value column
    sheet.cell(row=row_idx, column=3, value=row.duplicate_count)  # Count column

# Restore preserved red marks
for (old_row, col), (start_color, end_color) in existing_fills.items():
    cell_value = sheet.cell(row=old_row, column=col).value
    for new_row_idx in range(2, sheet.max_row + 1):
        if sheet.cell(row=new_row_idx, column=col).value == cell_value:
            sheet.cell(row=new_row_idx, column=col).fill = PatternFill(start_color=start_color, end_color=end_color, fill_type="solid")
            break

# Step 10: Set column widths
column_widths = [35, 15, 15]  # Define column widths for `name`, `value`, and `duplicate_count`
for idx, width in enumerate(column_widths, start=1):
    column_letter = get_column_letter(idx)
    sheet.column_dimensions[column_letter].width = width

# Save the updated workbook to athena_file
workbook.save(athena_file)

print(f"Excel file updated, formatted, and saved as '{athena_file}'.")
