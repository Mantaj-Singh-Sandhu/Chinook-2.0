import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment


def extract_output_sheet_name(input_file):
    """
    Reads the input file and extracts 5 digits starting after the 7th character.
    """
    with open(input_file, 'r') as file:
        content = file.readline().strip()
        if len(content) >= 13:  # Ensure there are at least 12 characters
            return content[6:13]  # Extract exactly 5 characters starting from the 8th character
        else:
            raise ValueError("Input text is too short to extract the sheet name.")


def get_output_file_name(devices_list_file, date_file):
    """
    Constructs the output file name using the first line from devices_list.txt
    and the date in date.txt.
    """
    # Read the first line of devices_list.txt
    with open(devices_list_file, 'r') as devices_file:
        device_name = devices_file.readline().strip()

    # Read the date from date.txt and format it
    with open(date_file, 'r') as date_file:
        date_str = date_file.readline().strip()
        try:
            formatted_date = datetime.strptime(date_str, "%Y-%m-%d").strftime("%Y%m%d")
        except ValueError:
            raise ValueError("Date in date.txt must be in YYYY-MM-DD format.")

    # Combine device name and date for the output file name
    return f"Surprise/{device_name}_{formatted_date}.xlsx", device_name, formatted_date


def copy_and_paste_excel(file_list, headings, output_sheet_name, output_file, folder_path="excel_outputs", spacing=5):
    """
    Copies content from multiple Excel files, adds custom headings, and pastes them into an output Excel file.
    :param file_list: List of filenames to process in a predefined sequence.
    :param headings: List of custom names for the headings.
    :param output_sheet_name: The name of the output sheet where data will be written.
    :param output_file: Path to the output Excel file.
    :param folder_path: Folder where the input files are stored.
    :param spacing: Number of rows to leave between pasted datasets.
    """
    if len(file_list) != len(headings):
        raise ValueError("The number of headings must match the number of files.")

    # Load or create the output workbook
    try:
        output_wb = openpyxl.load_workbook(output_file)
        print(f"Loaded existing workbook: {output_file}")
    except FileNotFoundError:
        output_wb = openpyxl.Workbook()
        print(f"Created a new workbook: {output_file}")

    # Ensure the output sheet exists
    if output_sheet_name not in output_wb.sheetnames:
        output_wb.create_sheet(output_sheet_name)
        print(f"Created output sheet: {output_sheet_name}")
    dest_sheet = output_wb[output_sheet_name]


    # Add date and device name to the first sheet (default sheet)
    first_sheet = output_wb.active
    first_sheet["M11"] = "Chinook"
    first_sheet["M13"] = device_name
    first_sheet["M14"] = formatted_date
    print(f"Added device name and date to the first sheet: {device_name}, {formatted_date}")

    # Define font styles and fill colors
    bold_font = Font(size=40, bold=True)  # Bold font with increased size
    heading_font = Font(size=16, bold=True)  # Bold font for headings
    blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Blue
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light Blue
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Lighter red
    center_alignment = Alignment(horizontal="center", vertical="center")  # Center alignment

    # Set styles for specific cells
    header_cells = ["M11"]
    for cell in header_cells:
        first_sheet[cell].font = bold_font
        first_sheet[cell].alignment = center_alignment

    # Apply fill colors
    first_sheet["M11"].fill = light_blue_fill  # Blue fill for "Chinook"

    # Adjust column widths
    columns_to_adjust = ["M", "N"]  # Adjusting columns M and N for better fit
    for column in columns_to_adjust:
        first_sheet.column_dimensions[column].width = 25  # Increase the width for better visibility

    # Apply heading font and alignment for clarity
    for cell in ["M13", "M14"]:
        first_sheet[cell].alignment = center_alignment

    # Set column widths for A, B, C, D, and E
    dest_sheet.column_dimensions['A'].width = 40
    dest_sheet.column_dimensions['B'].width = 20
    dest_sheet.column_dimensions['C'].width = 20
    dest_sheet.column_dimensions['D'].width = 20
    dest_sheet.column_dimensions['E'].width = 20

    for index, (input_file, heading) in enumerate(zip(file_list, headings), start=1):
        # Create the full path to the input file
        input_file_path = f"{folder_path}/{input_file}"

        # Open the input workbook
        try:
            input_wb = openpyxl.load_workbook(input_file_path, data_only=True)
            print(f"Opened input file: {input_file_path}")
        except FileNotFoundError:
            print(f"File not found: {input_file_path}. Skipping...")
            continue

        source_sheet_name = "Sheet1"  # Assumed input sheet name
        if source_sheet_name not in input_wb.sheetnames:
            print(f"Sheet '{source_sheet_name}' not found in '{input_file}'. Skipping...")
            continue

        source_sheet = input_wb[source_sheet_name]

        # Determine the range of cells with data in the source sheet
        min_row = source_sheet.min_row
        max_row = source_sheet.max_row
        min_col = source_sheet.min_column
        max_col = source_sheet.max_column

        # Find the next available row in the destination sheet
        if dest_sheet.max_row == 1 and dest_sheet.cell(1, 1).value is None:
            dest_start_row = 1
        else:
            dest_start_row = dest_sheet.max_row + spacing

        # Add custom heading with a background color and bold font
        heading_cell = dest_sheet.cell(row=dest_start_row, column=1, value=heading)

        # Set background color based on index (first 7 with red, remaining with light blue)
        if index <= 7:
            heading_cell.fill = red_fill
        else:
            heading_cell.fill = light_blue_fill

        heading_cell.font = heading_font  # Apply bold font and increased size

        # Copy content from source range to destination sheet
        for row_index, row in enumerate(
            source_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col), start=1):
            for col_index, cell in enumerate(row, start=1):
                dest_row = dest_start_row + row_index
                dest_col = col_index
                dest_sheet.cell(row=dest_row, column=dest_col, value=cell.value)

        # Apply light blue fill to the first row of data after the heading
        first_data_row = dest_start_row + 1  # The row immediately after the heading
        for col_index in range(1, max_col + 1):
            dest_sheet.cell(row=first_data_row, column=col_index).fill = light_blue_fill

    # Save the output workbook
    output_wb.save(output_file)
    print(f"Data successfully written to {output_file}")


# Specify the file list in the predefined sequence
file_list = [
    "athena_query_results_dtc_with_count.xlsx",
    "athena_query_results_error_no_duplicates.xlsx",
    "FMI-CID.xlsx",
    "J1939_non_duplicates.xlsx",
    "J1939_out_of_bounds.xlsx",
    "CDL_non_duplicates_file.xlsx",
    "CDL_out_of_bounds.xlsx",
    "merged_combined_statistics_ordered_CDL.xlsx",
    "merged_combined_statistics_ordered_J1939.xlsx",
]

# Define custom headings for each file
headings = [
    "DTC",
    "Error",
    "FMI-CID",
    "J1939 Stuck Tags",
    "J1939 Out of Range Tags",
    "CDL Stuck Tags",
    "CDL Out of Range Tags",
    "Combined CDL Statistics",
    "Combined J1939 Statistics",
]

# Define input file paths
input_file_path = "input_file.txt"  # For extracting sheet name
devices_list_file = "devices_list.txt"  # For the output file name
date_file = "date.txt"  # For appending the date to the output file name

# Read the output sheet name
output_sheet_name = extract_output_sheet_name(input_file_path)

# Generate the output file name and capture device name and formatted date
output_file, device_name, formatted_date = get_output_file_name(devices_list_file, date_file)

# Run the function
copy_and_paste_excel(file_list, headings, output_sheet_name, output_file, folder_path="excel_outputs", spacing=5)
